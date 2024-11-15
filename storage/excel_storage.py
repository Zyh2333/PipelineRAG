# storage/excel_storage.py
import os
from typing import Any, Dict, List, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from .base import BaseStorage


class ExcelStorage(BaseStorage):
    """Excel存储实现"""

    def __init__(self, file_path: str, sheet_name: str = "Sheet1"):
        """
        初始化Excel存储

        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
        """
        self.file_path = file_path
        self.sheet_name = sheet_name
        self._initialize_workbook()

    def _initialize_workbook(self):
        """初始化或加载工作簿"""
        if os.path.exists(self.file_path):
            try:
                self.workbook = load_workbook(self.file_path)
                if self.sheet_name not in self.workbook.sheetnames:
                    self.sheet = self.workbook.create_sheet(self.sheet_name)
                else:
                    self.sheet = self.workbook[self.sheet_name]
            except Exception as e:
                self.workbook = Workbook()
                self.sheet = self.workbook.active
                self.sheet.title = self.sheet_name
        else:
            # 创建新的工作簿
            os.makedirs(os.path.dirname(self.file_path), exist_ok=True)
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.title = self.sheet_name

        # 获取当前最大行
        self.current_row = self.sheet.max_row
        if self.current_row == 1:
            self._write_headers(["Timestamp", "Type", "Data"])

    def _write_headers(self, headers: List[str]):
        """写入表头"""
        for col, header in enumerate(headers, 1):
            self.sheet.cell(row=1, column=col, value=header)
        self.current_row = 2
        self._save()

    def _format_value(self, value: Any) -> str:
        """格式化值以便存储"""
        import numpy as np
        if isinstance(value, (dict, list)):
            import json

            def numpy_handler(obj):
                if isinstance(obj, np.ndarray):
                    return obj.tolist()
                return str(obj)

            return json.dumps(value, ensure_ascii=False, default=numpy_handler)
        elif isinstance(value, np.ndarray):
            return str(value.tolist())
        return str(value)

    def save(self, data: Any):
        """
        保存数据到Excel

        Args:
            data: 要保存的数据
        """
        from datetime import datetime

        # 写入时间戳
        self.sheet.cell(row=self.current_row, column=1,
                        value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

        # 写入数据类型
        self.sheet.cell(row=self.current_row, column=2,
                        value=type(data).__name__)

        # 写入数据
        formatted_data = self._format_value(data)
        self.sheet.cell(row=self.current_row, column=3,
                        value=formatted_data)

        self.current_row += 1
        self._save()

    def _save(self):
        """保存工作簿"""
        try:
            self.workbook.save(self.file_path)
        except Exception as e:
            print(f"保存Excel文件时出错: {str(e)}")

    def load(self) -> List[Dict]:
        """
        加载Excel中的数据

        Returns:
            List[Dict]: 加载的数据列表
        """
        data = []
        headers = [cell.value for cell in self.sheet[1]]

        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))
            try:
                import json
                row_data['Data'] = json.loads(row_data['Data'])
            except:
                pass
            data.append(row_data)

        return data

    def close(self):
        """关闭存储"""
        try:
            self.workbook.close()
        except:
            pass